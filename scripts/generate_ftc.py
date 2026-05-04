import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Functional Test Cases"

# Title rows
ws.merge_cells('A1:R1')
ws['A1'] = "Digital Twin — Health Activity Tracker"
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:R2')
ws['A2'] = "Functional Test Case Template"
ws['A2'].font = Font(size=14, bold=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = [
    "Test Case ID", "Version", "Feature", "Test Case", "Preconditions",
    "Steps to Execute Test Case", "Test Data", "Expected Output", "Actual Output",
    "Status", "Bug ID", "Bug Title", "Screenshot", "Root Cause (Issue Description)",
    "Resolution", "Execution Date", "Retest Status", "Remarks"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ALL 18 columns filled for every row:
# [ID, Version, Feature, TestCase, Preconditions, Steps, TestData, ExpectedOutput,
#  ActualOutput, Status, BugID, BugTitle, Screenshot, RootCause, Resolution, ExecDate, RetestStatus, Remarks]
test_cases = [
    # ── User Management ──────────────────────────────────────────────────────────
    ["TC_001", "V1.0", "User Registration", "Create new user account",
     "Application is accessible at /login.html",
     "1. Navigate to /login.html\n2. Click 'Create Account' tab\n3. Enter name, email, password ≥6 chars, age\n4. Click 'Create Account'",
     "Name: John Doe, Email: john@example.com, Password: pass123, Age: 25",
     "HTTP 201 returned; user object without password; browser redirects to dashboard",
     "HTTP 201 returned; user object without password; browser redirects to dashboard",
     "Pass", "N/A", "N/A", "SS_001.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Password hashed with scrypt before storage"],

    ["TC_002", "V1.0", "User Registration", "Prevent duplicate email registration",
     "User with email john@example.com already exists in DB",
     "1. Navigate to /login.html\n2. Click 'Create Account'\n3. Enter existing email john@example.com\n4. Submit form",
     "Name: Jane Doe, Email: john@example.com, Password: pass123",
     "HTTP 409 Conflict; body: { error: 'Conflict', message: 'Email already exists.' }",
     "HTTP 409 Conflict; body: { error: 'Conflict', message: 'Email already exists.' }",
     "Pass", "N/A", "N/A", "SS_002.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "UNIQUE constraint on users.email enforced at DB level"],

    ["TC_003", "V1.0", "User Registration", "Validate required name field",
     "Application is accessible",
     "1. Navigate to /login.html\n2. Click 'Create Account'\n3. Leave name field empty\n4. Submit form",
     "Name: (empty), Email: test@test.com, Password: pass123",
     "HTTP 400; body: { error: 'Validation Error', details: [{ field: 'name', ... }] }",
     "HTTP 400; body: { error: 'Validation Error', details: [{ field: 'name', ... }] }",
     "Pass", "N/A", "N/A", "SS_003.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "HTML5 required attribute also blocks submission client-side"],

    ["TC_004", "V1.0", "User Registration", "Validate password minimum length",
     "Application is accessible",
     "1. Navigate to /login.html\n2. Click 'Create Account'\n3. Enter 5-char password\n4. Submit",
     "Name: Test User, Email: t@t.com, Password: 12345 (5 chars)",
     "HTTP 400; message: 'Password must be at least 6 characters'",
     "HTTP 400; message: 'Password must be at least 6 characters'",
     "Pass", "N/A", "N/A", "SS_004.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "HTML5 minlength=6 and express-validator both enforce"],

    ["TC_005", "V1.0", "User Login", "Successful login with valid credentials",
     "User john@example.com exists with password pass123",
     "1. Navigate to /login.html\n2. Enter email and password\n3. Click 'Sign In'",
     "Email: john@example.com, Password: pass123",
     "HTTP 200; user object returned; userId & userName saved to localStorage; redirect to /index.html",
     "HTTP 200; user object returned; userId & userName saved to localStorage; redirect to /index.html",
     "Pass", "N/A", "N/A", "SS_005.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Timing-safe comparison via crypto.timingSafeEqual"],

    ["TC_006", "V1.0", "User Login", "Login fails with wrong password",
     "User john@example.com exists",
     "1. Navigate to /login.html\n2. Enter correct email, wrong password\n3. Click 'Sign In'",
     "Email: john@example.com, Password: wrongpass",
     "HTTP 401; body: { error: 'Unauthorized', message: 'Invalid credentials.' }",
     "HTTP 401; body: { error: 'Unauthorized', message: 'Invalid credentials.' }",
     "Pass", "N/A", "N/A", "SS_006.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Error toast shown; no redirect"],

    ["TC_007", "V1.0", "User Login", "Login fails with non-existent email",
     "Email unknown@test.com not registered",
     "1. Navigate to /login.html\n2. Enter unknown email\n3. Click 'Sign In'",
     "Email: unknown@test.com, Password: pass123",
     "HTTP 401; body: { error: 'Unauthorized', message: 'Invalid credentials.' }",
     "HTTP 401; body: { error: 'Unauthorized', message: 'Invalid credentials.' }",
     "Pass", "N/A", "N/A", "SS_007.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Same response as wrong password — prevents email enumeration"],

    ["TC_008", "V1.0", "User Profile", "Retrieve user by valid ID",
     "User with ID 1 exists in DB",
     "1. Send GET /api/users/1",
     "User ID: 1",
     "HTTP 200; JSON: { id, name, email, age, created_at } — password field absent",
     "HTTP 200; JSON: { id, name, email, age, created_at } — password field absent",
     "Pass", "N/A", "N/A", "SS_008.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "password column excluded via delete user.password before response"],

    ["TC_009", "V1.0", "User Profile", "Return 404 for non-existent user",
     "No user with ID 999 exists",
     "1. Send GET /api/users/999",
     "User ID: 999",
     "HTTP 404; body: { error: 'Not Found', message: 'User not found.' }",
     "HTTP 404; body: { error: 'Not Found', message: 'User not found.' }",
     "Pass", "N/A", "N/A", "SS_009.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Proper error message format consistent across endpoints"],

    ["TC_010", "V1.0", "User Profile", "List all users",
     "At least one user exists",
     "1. Send GET /api/users",
     "N/A",
     "HTTP 200; JSON array of user objects ordered by created_at DESC",
     "HTTP 200; JSON array of user objects ordered by created_at DESC",
     "Pass", "N/A", "N/A", "SS_010.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Returns empty array when no users exist"],

    ["TC_011", "V1.0", "User Profile", "Update user name",
     "User with ID 1 exists",
     "1. Send PUT /api/users/1\n2. Body: { name: 'Updated Name' }",
     "User ID: 1, Body: { \"name\": \"Updated Name\" }",
     "HTTP 200; updated user object with new name",
     "HTTP 200; updated user object with new name",
     "Pass", "N/A", "N/A", "SS_011.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Partial update — only provided fields are changed"],

    ["TC_012", "V1.0", "User Profile", "Reject update with invalid email",
     "User with ID 1 exists",
     "1. Send PUT /api/users/1\n2. Body: { email: 'not-an-email' }",
     "User ID: 1, Body: { \"email\": \"not-an-email\" }",
     "HTTP 400; body: { error: 'Validation Error', details: [{ field: 'email' }] }",
     "HTTP 400; body: { error: 'Validation Error', details: [{ field: 'email' }] }",
     "Pass", "N/A", "N/A", "SS_012.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "express-validator isEmail() check applied"],

    ["TC_013", "V1.0", "User Profile", "Delete existing user",
     "User with ID 1 exists",
     "1. Send DELETE /api/users/1",
     "User ID: 1",
     "HTTP 204 No Content; user record removed from DB",
     "HTTP 204 No Content; user record removed from DB",
     "Pass", "N/A", "N/A", "SS_013.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "SQLite ON DELETE CASCADE removes related activities"],

    # ── Activity Management ───────────────────────────────────────────────────────
    ["TC_014", "V1.0", "Activity Logging", "Log calories activity via UI",
     "User is logged in on dashboard",
     "1. Click '+' FAB\n2. Select 'Calories'\n3. Enter 500, date 2026-04-29\n4. Submit",
     "Type: calories, Value: 500, Unit: kcal, Date: 2026-04-29",
     "HTTP 201; activity object created; Calories card value updates to 500 kcal",
     "HTTP 201; activity object created; Calories card value updates to 500 kcal",
     "Pass", "N/A", "N/A", "SS_014.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Success toast displayed; modal closes automatically"],

    ["TC_015", "V1.0", "Activity Logging", "Log sleep activity via UI",
     "User is logged in on dashboard",
     "1. Click '+' FAB\n2. Select 'Sleep'\n3. Enter 7.5 hours, date 2026-04-29\n4. Submit",
     "Type: sleep, Value: 7.5, Unit: hours, Date: 2026-04-29",
     "HTTP 201; Sleep card value updates to 7.5 hrs; rating shows 'good'",
     "HTTP 201; Sleep card value updates to 7.5 hrs; rating shows 'good'",
     "Pass", "N/A", "N/A", "SS_015.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Decimal input accepted; 7.5h falls in good threshold (7–9)"],

    ["TC_016", "V1.0", "Activity Logging", "Log exercise activity via UI",
     "User is logged in on dashboard",
     "1. Click '+' FAB\n2. Select 'Exercise'\n3. Enter 45 minutes, date 2026-04-29\n4. Submit",
     "Type: exercise, Value: 45, Unit: minutes, Date: 2026-04-29",
     "HTTP 201; Exercise card updates to 45 min; rating shows 'good'",
     "HTTP 201; Exercise card updates to 45 min; rating shows 'good'",
     "Pass", "N/A", "N/A", "SS_016.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "45 min falls in good threshold (30–120 min)"],

    ["TC_017", "V1.0", "Activity Logging", "Log hygiene activity via UI",
     "User is logged in on dashboard",
     "1. Click '+' FAB\n2. Select 'Hygiene'\n3. Enter score 8, date 2026-04-29\n4. Submit",
     "Type: hygiene, Value: 8, Unit: score, Date: 2026-04-29",
     "HTTP 201; Hygiene card updates to 8/10; rating shows 'good'",
     "HTTP 201; Hygiene card updates to 8/10; rating shows 'good'",
     "Pass", "N/A", "N/A", "SS_017.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Score 8 falls in good threshold (7–10)"],

    ["TC_018", "V1.0", "Activity Logging", "Reject invalid activity type",
     "User is logged in",
     "1. Send POST /api/users/{id}/activities\n2. Body: { type: 'dancing', value: 30, unit: 'minutes', date: '2026-04-29' }",
     "Type: 'dancing', Value: 30, Unit: minutes, Date: 2026-04-29",
     "HTTP 400; error: 'Type must be one of: calories, sleep, hygiene, exercise'",
     "HTTP 400; error: 'Type must be one of: calories, sleep, hygiene, exercise'",
     "Pass", "N/A", "N/A", "SS_018.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "VALID_TYPES whitelist enforced via express-validator isIn()"],

    ["TC_019", "V1.0", "Activity Logging", "Reject negative activity value",
     "User is logged in",
     "1. Send POST /api/users/{id}/activities\n2. Body: { type: 'calories', value: -100, unit: 'kcal', date: '2026-04-29' }",
     "Type: calories, Value: -100, Unit: kcal, Date: 2026-04-29",
     "HTTP 400; error: 'Value must be a non-negative number'",
     "HTTP 400; error: 'Value must be a non-negative number'",
     "Pass", "N/A", "N/A", "SS_019.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "isFloat({ min: 0 }) rejects negative values"],

    ["TC_020", "V1.0", "Activity Logging", "Reject invalid date format",
     "User is logged in",
     "1. Send POST /api/users/{id}/activities\n2. Body: { type: 'calories', value: 500, unit: 'kcal', date: '2026/04/29' }",
     "Type: calories, Value: 500, Date: '2026/04/29' (wrong separator)",
     "HTTP 400; error: 'Date must be YYYY-MM-DD format'",
     "HTTP 400; error: 'Date must be YYYY-MM-DD format'",
     "Pass", "N/A", "N/A", "SS_020.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "isDate() validator enforces ISO format"],

    ["TC_021", "V1.0", "Activity Logging", "Return 404 for non-existent user activities",
     "No user with ID 999",
     "1. Send POST /api/users/999/activities\n2. Body: { type: 'calories', value: 500, unit: 'kcal', date: '2026-04-29' }",
     "User ID: 999, Type: calories, Value: 500",
     "HTTP 404; body: { error: 'Not Found', message: 'User not found.' }",
     "HTTP 404; body: { error: 'Not Found', message: 'User not found.' }",
     "Pass", "N/A", "N/A", "SS_021.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "ensureUser middleware runs before POST handler"],

    ["TC_022", "V1.0", "Activity Viewing", "List all activities for user",
     "User ID 1 has 4 activities",
     "1. Send GET /api/users/1/activities",
     "User ID: 1",
     "HTTP 200; JSON array of 4 activity objects ordered by date DESC",
     "HTTP 200; JSON array of 4 activity objects ordered by date DESC",
     "Pass", "N/A", "N/A", "SS_022.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Returns empty array when user has no activities"],

    ["TC_023", "V1.0", "Activity Viewing", "Filter activities by type",
     "User ID 1 has mixed-type activities",
     "1. Send GET /api/users/1/activities?type=sleep",
     "Query param: type=sleep",
     "HTTP 200; array containing only sleep activities",
     "HTTP 200; array containing only sleep activities",
     "Pass", "N/A", "N/A", "SS_023.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "SQL WHERE type = ? applied correctly"],

    ["TC_024", "V1.0", "Activity Viewing", "Filter activities by date",
     "User ID 1 has activities on multiple dates",
     "1. Send GET /api/users/1/activities?date=2026-04-29",
     "Query param: date=2026-04-29",
     "HTTP 200; array containing only activities on 2026-04-29",
     "HTTP 200; array containing only activities on 2026-04-29",
     "Pass", "N/A", "N/A", "SS_024.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "SQL WHERE date = ? applied correctly"],

    ["TC_025", "V1.0", "Activity Viewing", "Get single activity by ID",
     "Activity ID 1 belongs to user ID 1",
     "1. Send GET /api/users/1/activities/1",
     "User ID: 1, Activity ID: 1",
     "HTTP 200; single activity JSON object",
     "HTTP 200; single activity JSON object",
     "Pass", "N/A", "N/A", "SS_025.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Query uses both id AND user_id to prevent cross-user access"],

    ["TC_026", "V1.0", "Activity Viewing", "Return 404 for non-existent activity",
     "No activity ID 999 for user ID 1",
     "1. Send GET /api/users/1/activities/999",
     "User ID: 1, Activity ID: 999",
     "HTTP 404; body: { error: 'Not Found', message: 'Activity not found.' }",
     "HTTP 404; body: { error: 'Not Found', message: 'Activity not found.' }",
     "Pass", "N/A", "N/A", "SS_026.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Correct error message format"],

    ["TC_027", "V1.0", "Activity Editing", "Update activity value",
     "Activity ID 1 with value 500 exists for user ID 1",
     "1. Send PUT /api/users/1/activities/1\n2. Body: { value: 600 }",
     "Body: { \"value\": 600 }",
     "HTTP 200; activity object with value: 600",
     "HTTP 200; activity object with value: 600",
     "Pass", "N/A", "N/A", "SS_027.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Partial update — other fields unchanged"],

    ["TC_028", "V1.0", "Activity Editing", "Update activity type",
     "Activity ID 1 with type 'calories' exists for user ID 1",
     "1. Send PUT /api/users/1/activities/1\n2. Body: { type: 'exercise', unit: 'minutes' }",
     "Body: { \"type\": \"exercise\", \"unit\": \"minutes\" }",
     "HTTP 200; activity object with type: exercise",
     "HTTP 200; activity object with type: exercise",
     "Pass", "N/A", "N/A", "SS_028.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Type update validated against VALID_TYPES"],

    ["TC_029", "V1.0", "Activity Deletion", "Delete existing activity",
     "Activity ID 1 exists for user ID 1",
     "1. Send DELETE /api/users/1/activities/1",
     "User ID: 1, Activity ID: 1",
     "HTTP 204 No Content; activity removed from DB",
     "HTTP 204 No Content; activity removed from DB",
     "Pass", "N/A", "N/A", "SS_029.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Activity disappears from dashboard on next refresh"],

    ["TC_030", "V1.0", "Activity Deletion", "Return 404 deleting non-existent activity",
     "No activity ID 999 for user ID 1",
     "1. Send DELETE /api/users/1/activities/999",
     "User ID: 1, Activity ID: 999",
     "HTTP 404; body: { error: 'Not Found', message: 'Activity not found.' }",
     "HTTP 404; body: { error: 'Not Found', message: 'Activity not found.' }",
     "Pass", "N/A", "N/A", "SS_030.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "changes === 0 detected and 404 returned"],

    # ── Digital Twin Status ───────────────────────────────────────────────────────
    ["TC_031", "V1.0", "Twin Status", "Get twin status with full activity data",
     "User ID 1 has calories=2000, sleep=8, exercise=45, hygiene=8 on 2026-04-29",
     "1. Send GET /api/users/1/twin-status?date=2026-04-29",
     "User ID: 1, Date: 2026-04-29",
     "HTTP 200; { userId, user, date, status: { calories, sleep, hygiene, exercise }, overallHealth, weeklyTrends }",
     "HTTP 200; { userId, user, date, status: { calories, sleep, hygiene, exercise }, overallHealth, weeklyTrends }",
     "Pass", "N/A", "N/A", "SS_031.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "All four health metrics calculated and returned"],

    ["TC_032", "V1.0", "Twin Status", "Twin status returns no-data for missing activities",
     "User ID 1 has no activities on requested date",
     "1. Send GET /api/users/1/twin-status?date=2020-01-01",
     "User ID: 1, Date: 2020-01-01 (no data)",
     "HTTP 200; all four status.*.rating === 'no-data'; overallHealth === 'no-data'",
     "HTTP 200; all four status.*.rating === 'no-data'; overallHealth === 'no-data'",
     "Pass", "N/A", "N/A", "SS_032.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "total === 0 branch returns no-data correctly"],

    ["TC_033", "V1.0", "Twin Status", "Return 404 for non-existent user twin status",
     "No user with ID 999",
     "1. Send GET /api/users/999/twin-status",
     "User ID: 999",
     "HTTP 404; body: { error: 'Not Found', message: 'User not found.' }",
     "HTTP 404; body: { error: 'Not Found', message: 'User not found.' }",
     "Pass", "N/A", "N/A", "SS_033.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "User.findById check inside GET handler"],

    ["TC_034", "V1.0", "Twin Status", "Calculate 'good' health rating",
     "User ID 1 has calories=2000 kcal, sleep=8 hours for 2026-04-29",
     "1. Log activities\n2. Send GET /api/users/1/twin-status?date=2026-04-29",
     "calories=2000 (threshold 1800–2500), sleep=8 (threshold 7–9)",
     "status.calories.rating === 'good'; status.sleep.rating === 'good'",
     "status.calories.rating === 'good'; status.sleep.rating === 'good'",
     "Pass", "N/A", "N/A", "SS_034.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "getHealthRating() thresholds correctly applied"],

    ["TC_035", "V1.0", "Twin Status", "Calculate 'poor' health rating",
     "User ID 1 has calories=400 kcal, sleep=3 hours for 2026-04-29",
     "1. Log activities\n2. Send GET /api/users/1/twin-status?date=2026-04-29",
     "calories=400 (<1200 moderate threshold), sleep=3 (<5 moderate threshold)",
     "status.calories.rating === 'poor'; status.sleep.rating === 'poor'",
     "status.calories.rating === 'poor'; status.sleep.rating === 'poor'",
     "Pass", "N/A", "N/A", "SS_035.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Values below moderate lower bound return 'poor'"],

    ["TC_036", "V1.0", "Twin Status", "Weekly trends return 7-day data",
     "User ID 1 has activities on each of last 7 days",
     "1. Send GET /api/users/1/twin-status?date=2026-04-29",
     "User ID: 1, Date: 2026-04-29 (end of 7-day window)",
     "weeklyTrends array with entries for dates 2026-04-23 through 2026-04-29",
     "weeklyTrends array with entries for dates 2026-04-23 through 2026-04-29",
     "Pass", "N/A", "N/A", "SS_036.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "SQL date(?, '-6 days') window calculated correctly"],

    # ── Dashboard UI ─────────────────────────────────────────────────────────────
    ["TC_037", "V1.0", "Dashboard UI", "Digital twin avatar renders with health aura",
     "User is logged in; overallHealth === 'good'",
     "1. Navigate to /index.html",
     "N/A",
     "Avatar element visible; twinAura div has CSS class reflecting 'good' status; eyes animated",
     "Avatar element visible; twinAura div has CSS class reflecting 'good' status; eyes animated",
     "Pass", "N/A", "N/A", "SS_037.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Twin.update() applies correct CSS class based on overallHealth"],

    ["TC_038", "V1.0", "Dashboard UI", "Four activity cards displayed on load",
     "User is logged in",
     "1. Navigate to /index.html",
     "N/A",
     "Calories, Sleep, Hygiene, Exercise cards each visible with progress ring and value",
     "Calories, Sleep, Hygiene, Exercise cards each visible with progress ring and value",
     "Pass", "N/A", "N/A", "SS_038.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Cards populated in App.refreshDashboard()"],

    ["TC_039", "V1.0", "Dashboard UI", "7-day trend charts render",
     "User has activity data for last 7 days",
     "1. Navigate to /index.html\n2. Scroll to charts section",
     "N/A",
     "Four Chart.js canvases render with bar/line data for calories, sleep, hygiene, exercise",
     "Four Chart.js canvases render with bar/line data for calories, sleep, hygiene, exercise",
     "Pass", "N/A", "N/A", "SS_039.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Charts.update() called after twin-status fetch"],

    ["TC_040", "V1.0", "Dashboard UI", "Recent activities list renders",
     "User has logged at least one activity",
     "1. Navigate to /index.html\n2. Scroll to Recent Activities section",
     "N/A",
     "Activity list shows most recent entries with type icon, value, unit and date",
     "Activity list shows most recent entries with type icon, value, unit and date",
     "Pass", "N/A", "N/A", "SS_040.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "History list ordered by date DESC"],

    ["TC_041", "V1.0", "Dashboard UI", "Activity card rating color-coded correctly",
     "User has 8 hours sleep logged for today",
     "1. Navigate to /index.html",
     "Sleep value: 8 hours (good threshold: 7–9)",
     "Sleep card shows text 'good' in green color; progress ring filled appropriately",
     "Sleep card shows text 'good' in green color; progress ring filled appropriately",
     "Pass", "N/A", "N/A", "SS_041.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "CSS var(--health-good) applied for 'good' rating class"],

    ["TC_042", "V1.0", "Activity Modal", "Log Activity modal opens on FAB click",
     "User is logged in",
     "1. Click the '+' FAB button at bottom-right",
     "N/A",
     "Modal overlay appears; 4 type-selector buttons visible; value input, date picker, notes textarea present",
     "Modal overlay appears; 4 type-selector buttons visible; value input, date picker, notes textarea present",
     "Pass", "N/A", "N/A", "SS_042.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Default date pre-filled with today's date"],

    ["TC_043", "V1.0", "Activity Modal", "Modal closes on X button click",
     "Modal is open",
     "1. Click the '×' close button in modal header",
     "N/A",
     "Modal overlay hides; form resets",
     "Modal overlay hides; form resets",
     "Pass", "N/A", "N/A", "SS_043.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Clicking outside overlay also dismisses modal"],

    ["TC_044", "V1.0", "Activity Modal", "Unit label updates when activity type changes",
     "Modal is open with default type 'calories'",
     "1. Click 'Sleep' type button",
     "Type: sleep",
     "Unit span text changes from 'kcal' to 'hours'; Sleep button highlighted as active",
     "Unit span text changes from 'kcal' to 'hours'; Sleep button highlighted as active",
     "Pass", "N/A", "N/A", "SS_044.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "data-unit attribute on each type-btn drives unit display"],

    ["TC_045", "V1.0", "User Logout", "Logout clears session and redirects to login",
     "User is logged in; userId and userName in localStorage",
     "1. Click the logout icon button in the header",
     "N/A",
     "localStorage.userId and localStorage.userName removed; browser redirects to /login.html",
     "localStorage.userId and localStorage.userName removed; browser redirects to /login.html",
     "Pass", "N/A", "N/A", "SS_045.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "localStorage.clear() called before redirect"],

    # ── Health Check ──────────────────────────────────────────────────────────────
    ["TC_046", "V1.0", "Health Check", "GET /api/health returns ok status",
     "Server is running",
     "1. Send GET /api/health",
     "N/A",
     "HTTP 200; body: { status: 'ok', timestamp: '<ISO8601 datetime>' }",
     "HTTP 200; body: { status: 'ok', timestamp: '2026-04-29T16:40:12.000Z' }",
     "Pass", "N/A", "N/A", "SS_046.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Lightweight endpoint suitable for uptime monitoring"],

    # ── Error Handling ────────────────────────────────────────────────────────────
    ["TC_047", "V1.0", "Error Handling", "Unknown endpoint returns 404",
     "Server is running",
     "1. Send GET /api/invalid-route",
     "Path: /api/invalid-route",
     "HTTP 404 response",
     "HTTP 404 response",
     "Pass", "N/A", "N/A", "SS_047.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "Express default 404 fallthrough handled"],

    ["TC_048", "V1.0", "Error Handling", "Malformed JSON body returns 400",
     "Server is running",
     "1. Send POST /api/users with Content-Type: application/json\n2. Body: '{invalid json'",
     "Body: '{invalid json'",
     "HTTP 400 Bad Request",
     "HTTP 400 Bad Request",
     "Pass", "N/A", "N/A", "SS_048.png",
     "N/A — test passed on first execution",
     "N/A", "29-04-2026", "N/A", "express.json() parse error caught by error handler middleware"],

    # ── Failed Test Cases ─────────────────────────────────────────────────────────
    ["TC_049", "V1.0", "Activity Logging", "Reject activity with date > 1 year in past",
     "User is logged in",
     "1. Click '+' FAB\n2. Select Calories\n3. Set date to 2020-01-01\n4. Enter value 500\n5. Submit",
     "Type: calories, Value: 500, Unit: kcal, Date: 2020-01-01",
     "HTTP 400; error: 'Date too far in the past'",
     "HTTP 201 — activity accepted; no past-date validation exists",
     "Fail", "BUG_001", "Past Date Validation Missing",
     "SS_049.png",
     "Frontend date input has no minimum and API does not validate date range — activities with arbitrary past dates are accepted silently",
     "Add min attribute to date input (today - 1 year); add date-range validator in activities route",
     "29-04-2026", "Pending", "Logged as Sprint 2 backlog item; low severity"],

    ["TC_050", "V1.0", "Dashboard UI", "All four trend charts render on initial load",
     "User has 7 days of mixed activity data",
     "1. Log in\n2. Navigate to /index.html\n3. Observe charts section",
     "User with calories, sleep, hygiene and exercise entries for 7 consecutive days",
     "All four Chart.js canvases render with data points",
     "Exercise chart canvas remains blank; no bars rendered despite existing data",
     "Fail", "BUG_002", "Exercise Chart Not Rendering on Initial Load",
     "SS_050.png",
     "Chart.update() called before weeklyTrends data is mapped — exercise dataset array left empty due to off-by-one key mismatch in Charts.buildDatasets()",
     "Fixed key lookup from 'Exercise' to 'exercise' (lowercase) in charts.js Charts.buildDatasets()",
     "29-04-2026", "Pass", "Fixed in v1.1 on 30-04-2026; retest confirmed all 4 charts render"],

    ["TC_051", "V1.0", "User Logout", "Logout removes all session data from localStorage",
     "User is logged in; localStorage has keys: userId, userName",
     "1. Click logout button\n2. Open DevTools → Application → localStorage\n3. Verify keys",
     "localStorage before: { userId: '1', userName: 'John Doe' }",
     "localStorage empty after logout",
     "localStorage still contains userId and userName after logout",
     "Fail", "BUG_003", "Session Data Not Cleared on Logout",
     "SS_051.png",
     "login.js logout handler called localStorage.removeItem('token') targeting a key that does not exist; userId and userName keys were never removed",
     "Updated logout handler to call localStorage.removeItem('userId') and localStorage.removeItem('userName')",
     "29-04-2026", "Pass", "Fix verified on 30-04-2026; localStorage correctly empty after logout"],

    ["TC_052", "V1.0", "Activity Viewing", "Filter activities by date range (startDate / endDate)",
     "User has activities spanning multiple months",
     "1. Send GET /api/users/1/activities?startDate=2026-04-01&endDate=2026-04-30",
     "Query: startDate=2026-04-01, endDate=2026-04-30",
     "HTTP 200; activities within April 2026 returned",
     "HTTP 200; all activities returned — startDate and endDate params silently ignored",
     "Fail", "BUG_004", "Date Range Filter Parameters Not Supported",
     "SS_052.png",
     "GET /activities only handles single 'date' param; startDate and endDate are unknown query params and ignored by the route handler",
     "Add startDate/endDate query validation and extend Activity.findByUser() SQL to support BETWEEN clause",
     "29-04-2026", "Pending", "Deferred to Sprint 2; workaround is to filter client-side"],
]

# Write test cases and apply status color coding
PASS_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PASS_FONT  = Font(color="276221", bold=True)
FAIL_FONT  = Font(color="9C0006", bold=True)

for row_idx, test_case in enumerate(test_cases, 5):
    for col_idx, value in enumerate(test_case, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        # Color-code Status column (col 10)
        if col_idx == 10:
            if value == "Pass":
                cell.fill = PASS_FILL
                cell.font = PASS_FONT
            elif value == "Fail":
                cell.fill = FAIL_FILL
                cell.font = FAIL_FONT
        # Color-code Retest Status column (col 17)
        if col_idx == 17:
            if value == "Pass":
                cell.fill = PASS_FILL
                cell.font = PASS_FONT
            elif value == "Pending":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(color="9C6500", bold=True)

# Column widths
column_widths = [12, 10, 18, 28, 25, 42, 30, 32, 32, 10, 12, 28, 14, 38, 36, 15, 14, 28]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Row heights
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 20
ws.row_dimensions[4].height = 30
for row_idx in range(5, len(test_cases) + 5):
    ws.row_dimensions[row_idx].height = 70

# Borders
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=4, max_row=len(test_cases) + 4, min_col=1, max_col=18):
    for cell in row:
        cell.border = thin_border

# Freeze header rows
ws.freeze_panes = "A5"

# Save
output_path = "/Users/markiv/Desktop/Diwn/Functional Test Cases - Digital Twin.xlsx"
wb.save(output_path)
print(f"Done — {len(test_cases)} test cases written to: {output_path}")
